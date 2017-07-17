#!/usr/bin/env python
"""Auxiliary program for write in spreadsheet with format

It contains functions that write the poses data and time in a
spreadsheet with format. Some of the results are written dynamically,
allowing their modification by altering certain values.

It also contains a function to read data from an excel sheet and pass
it to a NumPy (float64) array.
"""
# Standard libraries
import numpy as np
# Excel read/write library
import openpyxl

try:
    # uvispace settings.
    import settings
except ImportError:
    # Exit program if the settings module can't be found.
    sys.exit("Can't find settings module. Maybe environment variables are not"
             "set. Run the environment .sh script at the project root folder.")

# Dictionary with different formats to apply to the cells when writes
# the data in a spreadsheet.
cell_formats = {
    # Types of text alignment.
    'center_al': openpyxl.styles.Alignment(horizontal='center',
                                           vertical='center'),
    'right_al': openpyxl.styles.Alignment(horizontal='right',
                                          vertical='center'),
    'left_al': openpyxl.styles.Alignment(horizontal='left', vertical='center'),
    # Types of font.
    'title_ft': openpyxl.styles.Font(color='FFFFFFFF', size=20, bold=True),
    'white_ft': openpyxl.styles.Font(color='FFFFFFFF', bold=True),
    # Types of border.
    'thick_bd': openpyxl.styles.Border(
                                top=openpyxl.styles.Side(border_style='thick',
                                                         color='FF4143CA'),
                                bottom=openpyxl.styles.Side(
                                       border_style='thick', color='FF4143CA')),
    'thin_bd': openpyxl.styles.Border(
                                   top=openpyxl.styles.Side(border_style='thin',
                                                            color='FF83C6D6'),
                                bottom=openpyxl.styles.Side(border_style='thin',
                                                            color='FF83C6D6')),
    # Types of fill.
    'blue_fill': openpyxl.styles.PatternFill(fill_type='solid',
                                             start_color='FF2F79E6',
                                             end_color='FF2F79E6'),
    'skyblue_fill': openpyxl.styles.PatternFill(fill_type='solid',
                                                start_color='FFDBF4FA',
                                                end_color='FFDBF4FA'),
    'white_fill': openpyxl.styles.PatternFill(fill_type='solid',
                                              start_color='FFFFFFFF',
                                              end_color='FFFFFFFF')
}


def read_data(spreadsheet_name):
    """It reads poses and time from spreadsheet and returns them back.

    Each group of data is composed by 4 elements with the location of
    the UGV at a given sample: time, "x" position, "y" position and
    angle "theta".

    :param str spreadsheet_name: name of spreadsheet that contain the
    data to be read.
    :return: Mx4 matrix with read data. M corresponds to the number of
    data samples (rows).
    :rtype: numpy.array(shape=Mx4)
    """
    # Open spreadsheet
    try:
        wb = openpyxl.load_workbook(spreadsheet_name)
    except IOError:
        wb = openpyxl.Workbook()
    ws = wb.active
    # Initialization of matrixes for data.
    new_data = np.array([0, 0, 0, 0]).astype(np.float64)
    # Rows 1 and 2 are for the file name, the next three rows describe the
    # experiment parameters, and the 6 is the header in any type file.
    # Start reading in row 7.
    row = 7
    # Loop for reading data.
    data_in_this_row = True
    while data_in_this_row:
        element = ws.cell(column=1, row=row).value
        # "Sum differerential data:" is the value of the next row to the last
        # data.
        if not element:
            data_in_this_row = False
        else:
            # Reading data.
            for col in range(4):
                new_data[col] = ws.cell(column=col+1, row=row).value
            if row == 7:
                # Substitution of row of zeros by first row of data read.
                data = new_data
            else:
                data = np.vstack([data, new_data])
            row +=1
    rounded_data = np.round(data, 2)
    return rounded_data


def write_spreadsheet(header, data_to_save, exp_name, exp_conditions,
                      save_analyzed):
    """
    Receive poses and time; and format and write them to a spreadsheet.

    :param header: contains the header to save in spreadsheet. It can
    have 4 or 11 elements (M).
    :type header: numpy.array str (1xM)
    :param data_to_save: contains data to save in spreadsheet.
    :param str exp_name: name of experiment where the data will
     be saved.
    :param str exp_conditions: contains string with experiment
     description.
    :param save_analyzed: boolean for save or not data in master file.
    :return: file name where data was saved.
    :rtype: str
    """
    spreadsheet_name = "datatemp/{}.xlsx".format(exp_name)
    try:
        wb = openpyxl.load_workbook(spreadsheet_name)
    except IOError:
        wb = openpyxl.Workbook()
    ws = wb.active
    # Spreadsheet title.
    ws.merge_cells('A1:K2')
    ws.cell('A1').value = exp_name
    ws.cell('A1').alignment = cell_formats['center_al']
    ws.cell('A1').font = cell_formats['title_ft']
    ws.cell('A1').fill = cell_formats['blue_fill']
    # Experiment conditions.
    ws.merge_cells('A3:K5')
    ws.cell('A3').value = exp_conditions
    # Freeze header
    ws.freeze_panes = ws['B7']
    # Write in the spreadsheet the header.
    rows, cols = data_to_save.shape
    for col in range(cols):
        ws.cell(column=col+1, row=6, value=header[col])
        ws.cell(column=col+1, row=6).alignment = cell_formats['right_al']
        ws.cell(column=col+1, row=6).font = cell_formats['white_ft']
        ws.cell(column=col+1, row=6).fill = cell_formats['blue_fill']
        ws.cell(column=col+1, row=6).border = cell_formats['thick_bd']
        my_cell = ws.cell(column=col+1, row=6)
        ws.column_dimensions[my_cell.column].width = 10
    # Write in the spreadsheet the data.
    for row in range(rows):
        for col in range(cols):
            element = float(data_to_save[row, col])
            cell = ws.cell(column=col+1, row=row+7, value=element)
            cell.number_format = '0.00'
            # Different fill color to consecutive rows.
            if row % 2 != 0:
                ws.cell(column=col+1, row=row+7).fill = cell_formats[
                                                                 'skyblue_fill']
            else:
                ws.cell(column=col+1, row=row+7).fill = cell_formats[
                                                                   'white_fill']
    # Writing and calculating statistics only if save_analyzed is True.
    if save_analyzed:
        # Headers of statistics.
        ws.cell(column=2, row=rows+7, value='Sum differential data:')
        ws.cell(column=2, row=rows+8, value='Mean differential data:')
        ws.cell(column=2, row=rows+9, value='Variance differential data:')
        ws.cell(column=2, row=rows+10, value='Std deviation differential data:')
        # Write in spreadsheet headers of average speeds.
        ws.cell(column=8, row=rows+11, value='Average Linear Speed:')
        ws.cell(column=8, row=rows+12, value='Average Angular Speed:')
        # Format to average speeds.
        ws.merge_cells(start_row=rows+11,start_column=8,end_row=rows+11,
                       end_column=10)
        ws.merge_cells(start_row=rows+12,start_column=8,end_row=rows+12,
                       end_column=10)
        # Write to spreadsheet statistics and average speeds.
        for row in range(rows+7, rows+13):
            #Format to statistics average speeds.
            ws.merge_cells(start_row=row, start_column=2, end_row=row,
                           end_column=4)
            # Write and calculate in spreadsheet the dynamic statistics.
            for col in range(1, cols+1):
                if col >= 5:
                    letter_range = openpyxl.utils.get_column_letter(col)
                    start_range = '{}{}'.format(letter_range, 7)
                    end_range = '{}{}'.format(letter_range, rows+6)
                    interval = '{}:{}'.format(start_range,end_range)
                    ws.cell(column=col, row=rows+7,
                            value= '=SUM({})\n'.format(interval))
                    ws.cell(column=col, row=rows+8,
                            value= '=AVERAGE({})\n'.format(interval))
                    ws.cell(column=col, row=rows+9,
                            value= '=VAR({})\n'.format(interval))
                    ws.cell(column=col, row=rows+10,
                            value= '=STDEV({})\n'.format(interval))
                # Format to dynamic statistics.
                cell = ws.cell(column=col, row=row)
                cell.number_format = '0.00'
                cell.font = cell_formats['white_ft']
                cell.fill = cell_formats['blue_fill']
                cell.alignment = cell_formats['right_al']
        # Dynamic average speeds.
        ws.cell(column=11, row=rows+11,
                value= '=1000*SQRT(((B{finalrow}-B7)^2)'
                        '+(((C{finalrow}-C7)^2)))/E{rowtime}\n'
                        ''.format(finalrow=rows+6, rowtime=rows+7))
        ws.cell(column=11, row=rows+12,
                value= '=1000*(D{lastdata}-D7)/E{rowtime}\n'
                       ''.format(lastdata=rows+6, rowtime=rows+7))

    wb.save(spreadsheet_name)
    return spreadsheet_name


def save2master_xlsx(exp_name, sp_left, sp_right, avg_lin_spd, avg_ang_spd):
    """Save absolute linear and angular speeds in a master spreadsheet.

    :param str exp_name: name of experiment realized.
    :param int sp_left: left speed setpoint.
    :param int sp_right: right speed setpoint.
    :param float64 avg_lin_spd: average linear speed of UGV.
    :param float64 avg_ang_spd: average angular speed of UGV.
    """
    # Data search to save in masterspreadsheet.
    master_dir = "'file://{}{}".format(settings.parent_path,
                                      "/uvisensor/datatemp/")
    spreadsheet_ext= '.xlsx'
    name_sheet = '\'#$Sheet.'
    avg_spd_column = 'K'
    try:
        wb = openpyxl.load_workbook('datatemp/{}.xlsx'.format(exp_name))
    except IOError:
        wb = openpyxl.Workbook()
    ws = wb.active
    # Search the next empty row.
    row = 7
    written_row = True
    while written_row:
        element = ws.cell(column=1, row=row).value
        if element is None:
            written_row = False
        else:
            row += 1
    dynamic_data = ''.join([master_dir, exp_name, spreadsheet_ext, name_sheet,
                            avg_spd_column])
    dynamic_avg_lin_spd = '{}{}'.format(dynamic_data, row+4)
    dynamic_avg_avg_spd = '{}{}'.format(dynamic_data, (row+5))
    # Save data in masterspreadsheet.
    try:
        wb = openpyxl.load_workbook("datatemp/masterfile.xlsx")
    except IOError:
        wb = openpyxl.Workbook()
    ws = wb.active
    # Freeze header
    ws.freeze_panes = ws['A2']
    # Search the next empty row.
    lastrow = 1
    written_row = True
    while written_row:
        element = ws.cell(column=1, row=lastrow).value
        if element is None:
            written_row = False
        else:
            lastrow += 1
    # Save data in empty row.
    ws.cell(column=1, row=lastrow, value=exp_name)
    ws.cell(column=2, row=lastrow, value=sp_left)
    ws.cell(column=3, row=lastrow, value=sp_right)
    # This data can be consulted without having the files of the experiments.
    ws.cell(column=4, row=lastrow, value=avg_lin_spd)
    ws.cell(column=5, row=lastrow, value=avg_ang_spd)
    # Indirect average spds data.
    ws.cell(column=6, row=lastrow,
            value= '=INDIRECT(H{})\n'.format(lastrow)).number_format = '0.00'
    ws.cell(column=7, row=lastrow,
            value= '=INDIRECT(I{})\n'.format(lastrow)).number_format = '0.00'
    # Address cell to read
    ws.cell(column=8, row=lastrow, value=dynamic_avg_lin_spd)
    ws.cell(column=9, row=lastrow, value=dynamic_avg_avg_spd)
    ws.cell(column=10, row=lastrow, value="_")
    #Format data.
    for col in range(1, 10):
        my_cell = ws.cell(column=col, row=lastrow)
        if col == 1:
            ws.column_dimensions[my_cell.column].width = 18
        else:
            ws.column_dimensions[my_cell.column].width = 10
        if lastrow % 2 != 0:
            ws.cell(column=col, row=lastrow).fill = cell_formats['skyblue_fill']
        else:
            ws.cell(column=col, row=lastrow).fill = cell_formats['white_fill']
        ws.cell(column=col, row=lastrow).alignment = cell_formats['right_al']
    wb.save("datatemp/masterfile.xlsx")
    return
